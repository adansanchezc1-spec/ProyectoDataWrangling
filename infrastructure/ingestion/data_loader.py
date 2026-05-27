"""Data loading utilities with optional pandas support.

Provides a `DataLoader` capable of loading CSV, JSON and Excel files.
If pandas is available it will be used for convenience; otherwise
fallbacks use stdlib parsers (csv/json) and openpyxl when available for Excel.
"""
from __future__ import annotations

from typing import Any, Dict, List, Optional
import os
import json
import csv

try:
    import pandas as pd
except Exception:  # pandas is optional
    pd = None


class DataLoader:
    def __init__(self, preview_rows: int = 5) -> None:
        self.preview_rows = preview_rows

    def load(self, path: str) -> Dict[str, Any]:
        """Load file at `path` and return a dict with keys: records, schema, preview.

        Supported formats: .csv, .json, .xlsx/.xls
        """
        if not os.path.exists(path):
            raise FileNotFoundError(path)

        _, ext = os.path.splitext(path.lower())
        if ext in (".csv",):
            return self._load_csv(path)
        if ext in (".json",):
            return self._load_json(path)
        if ext in (".xls", ".xlsx"):
            return self._load_excel(path)

        # fallback: try csv
        return self._load_csv(path)

    def _load_csv(self, path: str) -> Dict[str, Any]:
        if pd is not None:
            df = pd.read_csv(path)
            records = df.to_dict(orient="records")
            schema = {c: str(dtype) for c, dtype in zip(df.columns, df.dtypes)}
            preview = records[: self.preview_rows]
            return {"records": records, "schema": schema, "preview": preview}

        # stdlib csv fallback
        with open(path, newline="", encoding="utf-8") as fh:
            reader = csv.DictReader(fh)
            rows = [r for r in reader]
        schema = {k: "str" for k in (rows[0].keys() if rows else [])}
        return {"records": rows, "schema": schema, "preview": rows[: self.preview_rows]}

    def _load_json(self, path: str) -> Dict[str, Any]:
        with open(path, "r", encoding="utf-8") as fh:
            data = json.load(fh)

        # Accept either list of records or dict with key 'records'
        if isinstance(data, list):
            records = data
        elif isinstance(data, dict) and "records" in data and isinstance(data["records"], list):
            records = data["records"]
        else:
            # try to coerce dict-of-lists -> list of records
            if isinstance(data, dict):
                keys = list(data.keys())
                length = len(data[keys[0]]) if keys else 0
                records = [ {k: data[k][i] for k in keys} for i in range(length) ]
            else:
                records = []

        schema = {k: type(v).__name__ for k, v in (records[0].items() if records else [])}
        return {"records": records, "schema": schema, "preview": records[: self.preview_rows]}

    def _load_excel(self, path: str) -> Dict[str, Any]:
        if pd is None:
            try:
                # lightweight openpyxl read: only first sheet
                from openpyxl import load_workbook

                wb = load_workbook(path, read_only=True)
                ws = wb[wb.sheetnames[0]]
                rows = list(ws.values)
                if not rows:
                    return {"records": [], "schema": {}, "preview": []}
                headers = [str(h) for h in rows[0]]
                data_rows = rows[1:]
                records = []
                for r in data_rows:
                    records.append({h: v for h, v in zip(headers, r)})
                schema = {h: "str" for h in headers}
                return {"records": records, "schema": schema, "preview": records[: self.preview_rows]}
            except Exception as exc:
                raise RuntimeError("pandas or openpyxl required to read excel files") from exc

        df = pd.read_excel(path)
        records = df.to_dict(orient="records")
        schema = {c: str(dtype) for c, dtype in zip(df.columns, df.dtypes)}
        return {"records": records, "schema": schema, "preview": records[: self.preview_rows]}
